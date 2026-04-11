"""Project import and validation helpers for the control panel."""
from __future__ import annotations

import csv
import re
import uuid
from dataclasses import dataclass, field
from datetime import datetime, time
from pathlib import Path
from typing import Any

from aica.storage.adapters import normalize_group_alias
from aica.storage.contracts import ProjectRecord
from aica.text_sanitize import sanitize_text

try:  # pragma: no cover - import availability depends on runtime environment
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - fallback is validated by callers
    load_workbook = None


PROJECT_IMPORT_HEADERS = (
    "task_order_no",
    "project_name",
    "customer_name",
    "product_line",
    "product_version",
    "project_manager",
    "follow_up_started_at",
    "support_ended_at",
    "project_level",
    "group_aliases",
)

PROJECT_TEMPLATE_CONTENT = (
    "task_order_no,project_name,customer_name,product_line,product_version,"
    "project_manager,follow_up_started_at,support_ended_at,project_level,group_aliases\n"
    "WO-2026-001,示例项目,示例客户,客服中台,v3.1,张三,2026-01-01T09:00:00,"
    "2099-12-31T23:59:59,常规,\"示例客户群A;示例客户群B\"\n"
)


def _now_iso() -> str:
    return datetime.now().isoformat()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s\-]+", "_", sanitize_text(value).casefold())


def split_project_aliases(raw_value: Any) -> tuple[str, ...]:
    aliases: list[str] = []
    seen: set[str] = set()
    for item in re.split(r"[\n,，;；]+", sanitize_text(raw_value)):
        alias = sanitize_text(item)
        normalized = normalize_group_alias(alias)
        if not alias or not normalized or normalized in seen:
            continue
        aliases.append(alias)
        seen.add(normalized)
    return tuple(aliases)


def build_project_template_content() -> str:
    return PROJECT_TEMPLATE_CONTENT


def normalize_project_level(value: str) -> str:
    text = sanitize_text(value).casefold()
    if text in {"重要", "important", "high", "critical"}:
        return "important"
    return "normal"


def _parse_datetime_value(value: str, *, end_of_day: bool = False) -> datetime | None:
    text = sanitize_text(value)
    if not text:
        return None
    normalized = text.replace("/", "-")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        try:
            parsed_date = datetime.strptime(normalized, "%Y-%m-%d").date()
        except ValueError:
            return None
        parsed = datetime.combine(parsed_date, time.max if end_of_day else time.min)
    else:
        if "T" not in normalized and " " not in normalized:
            parsed = datetime.combine(parsed.date(), time.max if end_of_day else time.min)
    return parsed


def is_project_active(support_ended_at: str, *, now: str | None = None) -> bool:
    parsed_end = _parse_datetime_value(support_ended_at, end_of_day=True)
    if parsed_end is None:
        normalized_end = sanitize_text(support_ended_at)
        current_time = sanitize_text(now) or _now_iso()
        return not normalized_end or normalized_end >= current_time
    parsed_now = _parse_datetime_value(sanitize_text(now) or _now_iso()) or datetime.now()
    return parsed_end >= parsed_now


@dataclass(frozen=True)
class ProjectAliasConflict:
    row_number: int
    task_order_no: str
    alias: str
    conflicting_project_id: str
    conflicting_project_name: str
    conflicting_task_order_no: str

    def to_dict(self) -> dict[str, object]:
        return {
            "rowNumber": self.row_number,
            "taskOrderNo": self.task_order_no,
            "alias": self.alias,
            "conflictingProjectId": self.conflicting_project_id,
            "conflictingProjectName": self.conflicting_project_name,
            "conflictingTaskOrderNo": self.conflicting_task_order_no,
        }


@dataclass(frozen=True)
class ProjectImportRow:
    row_number: int
    task_order_no: str
    project_name: str
    customer_name: str = ""
    product_line: str = ""
    product_version: str = ""
    project_manager: str = ""
    follow_up_started_at: str = ""
    support_ended_at: str = ""
    project_level: str = "normal"
    aliases: tuple[str, ...] = ()


@dataclass
class ProjectImportResult:
    created_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    relinked_count: int = 0
    error_rows: list[dict[str, object]] = field(default_factory=list)
    alias_conflicts: list[ProjectAliasConflict] = field(default_factory=list)

    def to_dict(self) -> dict[str, object]:
        return {
            "createdCount": self.created_count,
            "updatedCount": self.updated_count,
            "skippedCount": self.skipped_count,
            "relinkedCount": self.relinked_count,
            "errorRows": list(self.error_rows),
            "aliasConflicts": [item.to_dict() for item in self.alias_conflicts],
        }


def _iter_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _iter_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("当前环境缺少 openpyxl，无法导入 .xlsx 文件")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:  # pragma: no branch - workbook should always be closed when opened
        workbook.close()
    if not rows:
        return []
    header_row = rows[0]
    headers = [_normalize_header(value) for value in header_row]
    payloads: list[dict[str, Any]] = []
    for row in rows[1:]:
        payloads.append({header: value for header, value in zip(headers, row) if header})
    return payloads


def load_project_import_rows(path: str | Path) -> tuple[list[ProjectImportRow], ProjectImportResult]:
    result = ProjectImportResult()
    target = Path(path)
    suffix = target.suffix.lower()
    if suffix == ".csv":
        raw_rows = _iter_csv_rows(target)
    elif suffix == ".xlsx":
        raw_rows = _iter_xlsx_rows(target)
    else:
        raise ValueError("仅支持导入 .csv 或 .xlsx 文件")

    rows: list[ProjectImportRow] = []
    for index, raw_row in enumerate(raw_rows, start=2):
        normalized_row = {_normalize_header(key): sanitize_text(value) for key, value in raw_row.items()}
        if not any(normalized_row.values()):
            result.skipped_count += 1
            continue
        task_order_no = normalized_row.get("task_order_no", "")
        project_name = normalized_row.get("project_name", "")
        if not task_order_no or not project_name:
            result.error_rows.append(
                {
                    "rowNumber": index,
                    "message": "task_order_no 和 project_name 为必填列",
                }
            )
            continue
        rows.append(
            ProjectImportRow(
                row_number=index,
                task_order_no=task_order_no,
                project_name=project_name,
                customer_name=normalized_row.get("customer_name", ""),
                product_line=normalized_row.get("product_line", ""),
                product_version=normalized_row.get("product_version", ""),
                project_manager=normalized_row.get("project_manager", ""),
                follow_up_started_at=normalized_row.get("follow_up_started_at", ""),
                support_ended_at=normalized_row.get("support_ended_at", ""),
                project_level=normalize_project_level(normalized_row.get("project_level", "") or "normal"),
                aliases=split_project_aliases(normalized_row.get("group_aliases", "")),
            )
        )
    return rows, result


def merge_project_record(existing: ProjectRecord | None, row: ProjectImportRow) -> ProjectRecord:
    return ProjectRecord(
        id=existing.id if existing is not None else str(uuid.uuid4()),
        project_name=row.project_name,
        customer_name=row.customer_name or (existing.customer_name if existing is not None else ""),
        task_order_no=row.task_order_no,
        follow_up_started_at=row.follow_up_started_at or (existing.follow_up_started_at if existing is not None else ""),
        support_ended_at=row.support_ended_at or (existing.support_ended_at if existing is not None else ""),
        product_line=row.product_line or (existing.product_line if existing is not None else ""),
        product_version=row.product_version or (existing.product_version if existing is not None else ""),
        project_manager=row.project_manager or (existing.project_manager if existing is not None else ""),
        project_level=normalize_project_level(
            row.project_level or (existing.project_level if existing is not None else "normal")
        ),
        aliases=row.aliases,
        created_at=existing.created_at if existing is not None else _now_iso(),
        updated_at=_now_iso(),
    )


def project_record_from_payload(payload: dict[str, Any], existing: ProjectRecord | None = None) -> ProjectRecord:
    task_order_no = sanitize_text(payload.get("taskOrderNo") or payload.get("task_order_no"))
    project_name = sanitize_text(payload.get("projectName") or payload.get("project_name"))
    if not task_order_no:
        raise ValueError("任务单号不能为空")
    if not project_name:
        raise ValueError("项目名称不能为空")
    aliases_payload = payload.get("aliases", ())
    aliases = split_project_aliases(
        "\n".join(str(item) for item in aliases_payload) if isinstance(aliases_payload, list) else aliases_payload
    )
    return ProjectRecord(
        id=sanitize_text(payload.get("id")) or (existing.id if existing is not None else str(uuid.uuid4())),
        project_name=project_name,
        customer_name=sanitize_text(payload.get("customerName") or payload.get("customer_name")),
        task_order_no=task_order_no,
        follow_up_started_at=sanitize_text(payload.get("followUpStartedAt") or payload.get("follow_up_started_at")),
        support_ended_at=sanitize_text(payload.get("supportEndedAt") or payload.get("support_ended_at")),
        product_line=sanitize_text(payload.get("productLine") or payload.get("product_line")),
        product_version=sanitize_text(payload.get("productVersion") or payload.get("product_version")),
        project_manager=sanitize_text(payload.get("projectManager") or payload.get("project_manager")),
        project_level=normalize_project_level(
            sanitize_text(payload.get("projectLevel") or payload.get("project_level")) or "normal"
        ),
        aliases=aliases,
        created_at=existing.created_at if existing is not None else _now_iso(),
        updated_at=_now_iso(),
    )


def project_to_payload(project: ProjectRecord, *, now: str | None = None) -> dict[str, object]:
    current_time = sanitize_text(now) or _now_iso()
    return {
        "id": project.id,
        "projectName": project.project_name,
        "customerName": project.customer_name,
        "taskOrderNo": project.task_order_no,
        "followUpStartedAt": project.follow_up_started_at,
        "supportEndedAt": project.support_ended_at,
        "productLine": project.product_line,
        "productVersion": project.product_version,
        "projectManager": project.project_manager,
        "projectLevel": project.project_level,
        "aliases": list(project.aliases),
        "aliasCount": len(project.aliases),
        "isExpired": not is_project_active(project.support_ended_at, now=current_time),
    }


def _release_project_aliases(
    reservations: dict[str, list[tuple[str, str, str]]],
    project_id: str,
) -> None:
    for normalized_alias in list(reservations.keys()):
        remaining = [item for item in reservations[normalized_alias] if item[0] != project_id]
        if remaining:
            reservations[normalized_alias] = remaining
            continue
        reservations.pop(normalized_alias, None)


def _reserve_project_aliases(
    reservations: dict[str, list[tuple[str, str, str]]],
    project: ProjectRecord,
    *,
    now: str,
) -> None:
    if not is_project_active(project.support_ended_at, now=now):
        return
    for alias in project.aliases:
        normalized_alias = normalize_group_alias(alias)
        if not normalized_alias:
            continue
        reservations.setdefault(normalized_alias, []).append(
            (project.id, project.project_name, project.task_order_no)
        )


def find_active_alias_conflicts(
    project: ProjectRecord,
    projects: list[ProjectRecord],
    *,
    now: str | None = None,
) -> list[ProjectAliasConflict]:
    current_time = sanitize_text(now) or _now_iso()
    conflicts: list[ProjectAliasConflict] = []
    reservations: dict[str, list[tuple[str, str, str]]] = {}
    for existing in projects:
        _reserve_project_aliases(reservations, existing, now=current_time)
    _release_project_aliases(reservations, project.id)
    if not is_project_active(project.support_ended_at, now=current_time):
        return []
    for alias in project.aliases:
        normalized_alias = normalize_group_alias(alias)
        for conflicting_id, conflicting_name, conflicting_task_order in reservations.get(normalized_alias, []):
            if conflicting_id == project.id:
                continue
            conflicts.append(
                ProjectAliasConflict(
                    row_number=0,
                    task_order_no=project.task_order_no,
                    alias=alias,
                    conflicting_project_id=conflicting_id,
                    conflicting_project_name=conflicting_name,
                    conflicting_task_order_no=conflicting_task_order,
                )
            )
    return conflicts


def import_projects_from_file(path: str | Path, project_repository, todo_store) -> ProjectImportResult:
    rows, result = load_project_import_rows(path)
    current_time = _now_iso()
    existing_projects = project_repository.list_projects(include_expired=True, now=current_time)
    existing_by_task_order = {project.task_order_no: project for project in existing_projects if project.task_order_no}
    reservations: dict[str, list[tuple[str, str, str]]] = {}
    for project in existing_projects:
        _reserve_project_aliases(reservations, project, now=current_time)

    saved_records: list[tuple[ProjectRecord, bool]] = []
    for row in rows:
        existing = existing_by_task_order.get(row.task_order_no)
        merged = merge_project_record(existing, row)
        _release_project_aliases(reservations, merged.id)
        row_conflicts: list[ProjectAliasConflict] = []
        if is_project_active(merged.support_ended_at, now=current_time):
            for alias in merged.aliases:
                normalized_alias = normalize_group_alias(alias)
                for conflict_id, conflict_name, conflict_task_order in reservations.get(normalized_alias, []):
                    if conflict_id == merged.id:
                        continue
                    row_conflicts.append(
                        ProjectAliasConflict(
                            row_number=row.row_number,
                            task_order_no=merged.task_order_no,
                            alias=alias,
                            conflicting_project_id=conflict_id,
                            conflicting_project_name=conflict_name,
                            conflicting_task_order_no=conflict_task_order,
                        )
                    )
        if row_conflicts:
            result.alias_conflicts.extend(row_conflicts)
            if existing is not None:
                _reserve_project_aliases(reservations, existing, now=current_time)
            continue
        saved_records.append((merged, existing is None))
        existing_by_task_order[merged.task_order_no] = merged
        _reserve_project_aliases(reservations, merged, now=current_time)

    for record, created in saved_records:
        project_repository.upsert_project(record)
        if created:
            result.created_count += 1
        else:
            result.updated_count += 1

    if saved_records:
        result.relinked_count = todo_store.relink_open_unresolved_todos()
    return result
